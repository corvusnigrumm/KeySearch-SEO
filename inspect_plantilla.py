from openpyxl import load_workbook

wb = load_workbook("PLANTILLA.xlsx")
print("Hojas:", wb.sheetnames)
print()

for sname in wb.sheetnames:
    ws = wb[sname]
    print(f"=== HOJA: {sname} ===")
    tab = ws.sheet_properties.tabColor
    print(f"  Tab color: {tab.rgb if tab else 'none'}")
    print(f"  Freeze panes: {ws.freeze_panes}")
    print(f"  Max row: {ws.max_row}, Max col: {ws.max_column}")

    for row_idx in range(1, min(8, ws.max_row + 1)):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            fill = cell.fill
            font = cell.font
            fill_color = "-"
            try:
                if fill and fill.fgColor and fill.fgColor.type == "rgb":
                    fill_color = fill.fgColor.rgb
            except Exception:
                pass
            font_color = "-"
            try:
                if font and font.color and font.color.type == "rgb":
                    font_color = font.color.rgb
            except Exception:
                pass
            val = repr(str(cell.value or ""))[:35]
            print(
                f"  [{row_idx},{col_idx}] val={val} bg={fill_color} fg={font_color} "
                f"bold={font.bold} size={font.size} name={font.name} align={cell.alignment.horizontal}"
            )

    print("  Anchos columnas:")
    for col_letter, col_dim in list(ws.column_dimensions.items())[:25]:
        print(f"    Col {col_letter}: width={col_dim.width}")

    print("  Altos filas 1-6:")
    for r in range(1, 7):
        print(f"    Fila {r}: height={ws.row_dimensions[r].height}")

    print()
