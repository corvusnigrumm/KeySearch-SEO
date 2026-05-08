from openpyxl import load_workbook

wb = load_workbook("PLANTILLA.xlsx")
ws = wb["Resumen"]

print("=== HOJA RESUMEN - primeras 22 filas ===")
for row_idx in range(1, 23):
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        fill = cell.fill
        font = cell.font
        fill_color = "-"
        try:
            if fill and fill.fgColor and fill.fgColor.type == "rgb":
                fill_color = fill.fgColor.rgb
        except Exception:
            pass
        font_color = "-"
        try:
            if font and font.color and font.color.type == "rgb":
                font_color = font.color.rgb
        except Exception:
            pass
        val = repr(str(cell.value or ""))[:60]
        if cell.value:
            print(
                f"  [{row_idx},{col_idx}] val={val} bg={fill_color} fg={font_color} "
                f"bold={font.bold} size={font.size} name={font.name} align={cell.alignment.horizontal}"
            )
