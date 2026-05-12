"""
Exportador de resultados a Excel.

El archivo prioriza trazabilidad y datos reales observables. Ya no muestra
rangos mensuales inventados.
"""
import os
from typing import Dict, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import APP_NAME, APP_VERSION, EXCEL_TEMPLATE_PATH, OUTPUT_DIR
from scraper.ai_filter import generar_bloques_editoriales
from scraper.utils import generar_nombre_archivo
from scraper.volume_estimator import ordenar_por_volumen

_DARK = "002060"
_DARK_ALT = "002060"
_ACCENT = "002060"
_RED = "C0392B"
_ORANGE = "E67E22"
_YELLOW = "F1C40F"
_GRAY = "95A5A6"
_WHITE = "FFFFFF"
_LIGHT_GRAY = "F7F9FC"
_BORDER_COLOR = "D5DDE5"
_GEO_FILL_COLOR = "E8F4FD"  # Azul muy claro para destacar columnas geo

FONT_HEADER = Font(name="Century Gothic", bold=True, size=11, color=_WHITE)
FONT_TITLE = Font(name="Century Gothic", bold=True, size=16, color=_DARK)
FONT_SUBTITLE = Font(name="Century Gothic", size=11, color=_GRAY)
FONT_BODY = Font(name="Century Gothic", size=10, color="333333")
FONT_BODY_BOLD = Font(name="Century Gothic", bold=True, size=10, color="333333")
FONT_INDEX = Font(name="Century Gothic", size=10, color="333333")
FONT_LABEL = Font(name="Century Gothic", bold=True, size=10, color=_DARK)
FONT_STAT_VALUE = Font(name="Century Gothic", bold=True, size=12, color=_ACCENT)
FONT_SECTION = Font(name="Century Gothic", bold=True, size=11, color=_DARK)

FONT_SCORE_VERY_HIGH = Font(name="Century Gothic", bold=True, size=10, color=_RED)
FONT_SCORE_HIGH = Font(name="Century Gothic", bold=True, size=10, color=_ORANGE)
FONT_SCORE_MEDIUM = Font(name="Century Gothic", bold=True, size=10, color="7D6608")
FONT_SCORE_LOW = Font(name="Century Gothic", size=10, color=_GRAY)

FILL_SCORE_VERY_HIGH = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
FILL_SCORE_HIGH = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid")
FILL_SCORE_MEDIUM = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
FILL_SCORE_LOW = PatternFill(start_color="F2F4F4", end_color="F2F4F4", fill_type="solid")
FILL_HEADER = PatternFill(start_color=_DARK, end_color=_DARK, fill_type="solid")
FILL_HEADER_ALT = PatternFill(start_color=_DARK_ALT, end_color=_DARK_ALT, fill_type="solid")
FILL_EVEN = PatternFill(start_color=_LIGHT_GRAY, end_color=_LIGHT_GRAY, fill_type="solid")
FILL_ODD = PatternFill(start_color=_WHITE, end_color=_WHITE, fill_type="solid")
FILL_ACCENT_LIGHT = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
FILL_GEO = PatternFill(start_color=_GEO_FILL_COLOR, end_color=_GEO_FILL_COLOR, fill_type="solid")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

BORDER = Border(
    left=Side(style="thin", color=_BORDER_COLOR),
    right=Side(style="thin", color=_BORDER_COLOR),
    top=Side(style="thin", color=_BORDER_COLOR),
    bottom=Side(style="thin", color=_BORDER_COLOR),
)
BORDER_BOTTOM = Border(bottom=Side(style="medium", color=_DARK))


def _categorizar_score_texto(score: float) -> str:
    if score >= 80:
        return "Muy alta"
    if score >= 55:
        return "Alta"
    if score >= 30:
        return "Media"
    if score >= 15:
        return "Baja"
    return "Muy baja"


def _font_score(score: float) -> Font:
    if score >= 80:
        return FONT_SCORE_VERY_HIGH
    if score >= 55:
        return FONT_SCORE_HIGH
    if score >= 30:
        return FONT_SCORE_MEDIUM
    return FONT_SCORE_LOW


def _fill_score(score: float) -> PatternFill:
    if score >= 80:
        return FILL_SCORE_VERY_HIGH
    if score >= 55:
        return FILL_SCORE_HIGH
    if score >= 30:
        return FILL_SCORE_MEDIUM
    return FILL_SCORE_LOW


def _aplicar_celda(ws, row, col, value, font=FONT_BODY, alignment=ALIGN_LEFT, fill=None, border=BORDER):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    cell.alignment = alignment
    cell.border = border
    if fill:
        cell.fill = fill
    return cell


def _valor_trends(value):
    return "-" if value is None else value


def _geo_label(nombre_pais: str, codigo_pais: str) -> str:
    """Formatea la georeferenciacion como 'Colombia (CO)'."""
    nombre = nombre_pais or ""
    codigo = (codigo_pais or "").upper()
    if nombre and codigo:
        return f"{nombre} ({codigo})"
    return nombre or codigo or "-"


def _idioma_label(language_code: str) -> str:
    """Convierte codigo de idioma a nombre legible."""
    mapa = {
        "es": "Español",
        "en": "Ingles",
        "pt": "Portugues",
        "fr": "Frances",
        "de": "Aleman",
        "it": "Italiano",
    }
    return mapa.get((language_code or "").lower(), (language_code or "-").upper())


def _modo_reporte(datos: dict) -> str:
    google_ads = datos.get("google_ads", {}) or {}
    if google_ads.get("enabled") and google_ads.get("keywords_enriched", 0) > 0:
        return "Google Ads + Google Trends + senales observables"
    return "Senales observables + Google Trends"


def _estado_google_ads(datos: dict) -> str:
    google_ads = datos.get("google_ads", {}) or {}
    if google_ads.get("enabled") and google_ads.get("keywords_enriched", 0) > 0:
        return f"Activo ({google_ads.get('keywords_enriched', 0)} keywords)"
    return f"No disponible ({google_ads.get('reason', 'sin detalle')})"


def _crear_hoja_datos(wb: Workbook, titulo: str, items: List[str], volumenes: dict, encabezado_item: str, header_fill=FILL_HEADER, language_code: str = "es"):
    """Crea una hoja con trazabilidad real de cada termino."""
    sheet_exists = titulo in wb.sheetnames
    if sheet_exists:
        ws = wb[titulo]
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, 20):
                ws.cell(row=row_idx, column=col_idx).value = None
    else:
        ws = wb.create_sheet(title=titulo)

    encabezados = [
        "#",
        encabezado_item,
        "Pais (Geo)",
        "Idioma",
        "Categoria",
        "Subcategoria",
        "Referencia",
        "Fuente",
        "Pos.",
        "Ads avg/mes",
        "Ads comp.",
        "Ads idx",
        "Bid bajo",
        "Bid alto",
        "Score",
        "Prioridad",
        "GT prom. 12m",
        "GT pico",
        "GT ultimo",
    ]
    anchos = [6, 52, 20, 12, 20, 20, 26, 20, 8, 12, 12, 10, 14, 14, 10, 12, 14, 10, 10]

    if ws.max_row < 1 or ws.cell(row=1, column=1).value is None:
        for col_idx, (encabezado, ancho) in enumerate(zip(encabezados, anchos), 1):
            cell = ws.cell(row=1, column=col_idx, value=encabezado)
            cell.font = FONT_HEADER
            cell.fill = header_fill
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = ancho
    else:
        # Si la hoja ya existe (plantilla), solo nos aseguramos de que los encabezados coincidan
        # pero no alteramos el diseño ni los anchos de columna originales.
        for col_idx, encabezado in enumerate(encabezados, 1):
            if ws.cell(row=1, column=col_idx).value is None:
                ws.cell(row=1, column=col_idx, value=encabezado)

    estilos_base = {col: ws.cell(row=2, column=col) for col in range(1, 20)}

    items_ordenados = ordenar_por_volumen(items, volumenes)

    for row_idx, item in enumerate(items_ordenados, 2):
        vol = volumenes.get(item, {})
        score = vol.get("score", 0)
        prioridad = _categorizar_score_texto(score)
        row_fill = FILL_EVEN if row_idx % 2 == 0 else FILL_ODD

        geo_text = _geo_label(vol.get("pais", ""), vol.get("pais_codigo", ""))
        idioma_text = _idioma_label(language_code)

        valores = [
            row_idx - 1,
            item,
            geo_text,
            idioma_text,
            vol.get("categoria_padre", "-"),
            vol.get("subcategoria", "-"),
            vol.get("referencia", "-"),
            vol.get("fuente", "-"),
            vol.get("posicion_fuente", "-"),
            _valor_trends(vol.get("google_ads_avg_monthly_searches")),
            _valor_trends(vol.get("google_ads_competition")),
            _valor_trends(vol.get("google_ads_competition_index")),
            _valor_trends(vol.get("google_ads_low_top_of_page_bid_micros")),
            _valor_trends(vol.get("google_ads_high_top_of_page_bid_micros")),
            score,
            prioridad,
            _valor_trends(vol.get("google_trends_promedio")),
            _valor_trends(vol.get("google_trends_pico")),
            _valor_trends(vol.get("google_trends_ultimo")),
        ]

        for col_idx, value in enumerate(valores, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            base = estilos_base.get(col_idx)
            if base is not None and base.has_style:
                try:
                    # Intentamos copiar el estilo completo si es posible
                    cell._style = base._style
                except Exception:
                    # Si falla, al menos mantenemos el formato basico si no es nuevo
                    pass

        # Solo aplicamos recoloreo custom cuando NO trabajamos sobre la plantilla.
        if not sheet_exists:
            for col_idx in range(1, 20):
                cell = ws.cell(row=row_idx, column=col_idx)
                if col_idx == 15:
                    cell.font = _font_score(score)
                    cell.fill = _fill_score(score)
                elif col_idx == 16:
                    cell.font = _font_score(score)
                    cell.fill = row_fill
                else:
                    cell.fill = row_fill

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    ws.sheet_properties.tabColor = _DARK
    return ws


def _crear_hoja_resumen(wb: Workbook, keyword: str, datos: dict, volumenes: dict):
    """Crea una hoja resumen orientada a decisiones y trazabilidad."""
    sugerencias = datos.get("sugerencias", [])
    preguntas = datos.get("preguntas_paa", [])
    preguntas_ac = datos.get("preguntas_autocompletado", [])
    relacionadas = datos.get("busquedas_relacionadas", [])
    total = len(sugerencias) + len(preguntas) + len(preguntas_ac) + len(relacionadas)
    con_trends = sum(1 for item in volumenes.values() if item.get("google_trends_promedio") is not None)
    con_ads = sum(1 for item in volumenes.values() if item.get("google_ads_avg_monthly_searches") is not None)

    if "Resumen" in wb.sheetnames:
        ws = wb["Resumen"]
        
        # Escribir el título en A1 si existe la celda (manteniendo el estilo)
        ws.cell(row=1, column=1, value=f"Analisis de keyword: {keyword}")

        # Mapeo de estadísticas según la estructura de la plantilla (Filas 14-27)
        # Solo actualizamos los valores en la columna B para no tocar las etiquetas de la columna A
        mapping = {
            14: keyword,
            15: _geo_label(datos.get("country_name", ""), datos.get("country_code", "")),
            16: datos.get("country_code", "-").upper(),
            17: _idioma_label(datos.get("language_code", "es")),
            18: datos.get("category_name", "-"),
            19: datos.get("subcategory_name", "-"),
            20: _modo_reporte(datos),
            21: _estado_google_ads(datos),
            22: len(sugerencias),
            23: len(preguntas),
            24: len(preguntas_ac),
            25: len(relacionadas),
            26: con_ads,
            27: con_trends,
            28: total
        }

        for row_num, value in mapping.items():
            ws.cell(row=row_num, column=2, value=value)

        # Distribución por prioridad (Filas 33-37 en Columna B)
        categorias_orden = ["Muy alta", "Alta", "Media", "Baja", "Muy baja"]
        conteo = {categoria: 0 for categoria in categorias_orden}
        for metrica in volumenes.values():
            conteo[_categorizar_score_texto(metrica.get("score", 0))] += 1
        
        priority_rows = {
            "Muy alta": 33,
            "Alta": 34,
            "Media": 35,
            "Baja": 36,
            "Muy baja": 37
        }

        for cat in categorias_orden:
            row_num = priority_rows[cat]
            ws.cell(row=row_num, column=2, value=conteo[cat])

        # Bloques editoriales con IA (Groq/Llama) basados en Top 5 por fuente.
        top_autocomplete = ordenar_por_volumen(datos.get("sugerencias", []), volumenes)[:5]
        top_paa = ordenar_por_volumen(datos.get("preguntas_paa", []), volumenes)[:5]
        top_preguntas_ac = ordenar_por_volumen(datos.get("preguntas_autocompletado", []), volumenes)[:5]
        top_relacionadas = ordenar_por_volumen(datos.get("busquedas_relacionadas", []), volumenes)[:5]

        def _score(item: str):
            if not item:
                return ""
            value = volumenes.get(item, {}).get("score")
            if value is None:
                return ""
            try:
                return round(float(value), 1)
            except Exception:
                return value

        # Tabla TOP 5 en Resumen (filas 41-45). Solo actualizamos valores.
        # A:B Autocompletado | C:D PAA | E:F Preguntas AC | G:H slot extra | I:J Relacionadas
        for idx in range(5):
            row = 41 + idx
            auto_item = top_autocomplete[idx] if idx < len(top_autocomplete) else ""
            paa_item = top_paa[idx] if idx < len(top_paa) else ""
            preg_item = top_preguntas_ac[idx] if idx < len(top_preguntas_ac) else ""
            rel_item = top_relacionadas[idx] if idx < len(top_relacionadas) else ""

            ws.cell(row=row, column=1, value=auto_item)
            ws.cell(row=row, column=2, value=_score(auto_item))
            ws.cell(row=row, column=3, value=paa_item)
            ws.cell(row=row, column=4, value=_score(paa_item))
            ws.cell(row=row, column=5, value=preg_item)
            ws.cell(row=row, column=6, value=_score(preg_item))
            ws.cell(row=row, column=7, value="")  # limpiamos arrastre viejo
            ws.cell(row=row, column=8, value="")
            ws.cell(row=row, column=9, value=rel_item)
            ws.cell(row=row, column=10, value=_score(rel_item))

        all_items = list(volumenes.keys())
        top_keywords_trends = ordenar_por_volumen(all_items, volumenes)[:20]

        bloques = generar_bloques_editoriales(
            keyword_base=keyword,
            pais=datos.get("country_name", ""),
            top_autocomplete=top_autocomplete,
            top_paa=top_paa,
            top_preguntas_autocomplete=top_preguntas_ac,
            top_relacionadas=top_relacionadas,
            top_keywords_trends=top_keywords_trends,
        )

        row_ejes = 47
        row_propuesta = 54
        row_enfoque = 56
        row_titulos = 58
        row_subtitulos = 70
        row_keywords = 82

        for r in range(40, 120):
            val = str(ws.cell(row=r, column=1).value or "").strip()
            if not val:
                continue
            val_lower = val.lower()
            if "ejes estrat" in val_lower:
                row_ejes = r
            elif "propuesta" in val_lower:
                row_propuesta = r
            elif val_lower == "enfoque":
                row_enfoque = r
            elif val_lower == "títulos" or val_lower == "titulos":
                row_titulos = r
            elif "subtítulos" in val_lower or "subtitulos" in val_lower:
                row_subtitulos = r
            elif val_lower == "keywords":
                row_keywords = r

        for i in range(9):
            if i < len(bloques["ejes"]):
                ws.cell(row=row_ejes + 1 + i, column=1, value=bloques["ejes"][i])
        
        ws.cell(row=row_propuesta + 1, column=1, value=bloques["propuesta"])
        ws.cell(row=row_enfoque + 1, column=1, value=bloques["enfoque"])

        for i in range(10):
            if i < len(bloques["titulos"]):
                ws.cell(row=row_titulos + 1 + i, column=1, value=bloques["titulos"][i])
            if i < len(bloques["subtitulos"]):
                ws.cell(row=row_subtitulos + 1 + i, column=1, value=bloques["subtitulos"][i])
            if i < len(bloques["keywords_trends"]):
                ws.cell(row=row_keywords + 1 + i, column=1, value=bloques["keywords_trends"][i])

        return ws

    ws = wb.create_sheet(title="Resumen", index=0)

    ws.merge_cells("A1:D1")
    _aplicar_celda(ws, 1, 1, f"Analisis de keyword: {keyword}", font=FONT_TITLE, alignment=ALIGN_LEFT)
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:D2")
    _aplicar_celda(
        ws,
        2,
        1,
        f"Reporte generado por {APP_NAME} V {APP_VERSION}",
        font=FONT_SUBTITLE,
        alignment=ALIGN_LEFT,
    )
    ws.row_dimensions[2].height = 20

    for col in range(1, 5):
        ws.cell(row=3, column=col).border = BORDER_BOTTOM

    row = 5
    ws.merge_cells(f"A{row}:D{row}")
    _aplicar_celda(ws, row, 1, "RESUMEN DE RESULTADOS", font=FONT_HEADER, alignment=ALIGN_CENTER, fill=FILL_HEADER)
    for col in range(2, 5):
        ws.cell(row=row, column=col).border = BORDER
        ws.cell(row=row, column=col).fill = FILL_HEADER
    row += 1

    stats = [
        ("Keyword analizada", keyword, None),
        ("Pais analizado", _geo_label(datos.get("country_name", ""), datos.get("country_code", "")), None),
        ("Codigo de pais", datos.get("country_code", "-").upper(), None),
        ("Idioma", _idioma_label(datos.get("language_code", "es")), None),
        ("Categoria", datos.get("category_name", "-"), None),
        ("Subcategoria", datos.get("subcategory_name", "-"), None),
        ("Modo del reporte", _modo_reporte(datos), None),
        ("Estado Google Ads", _estado_google_ads(datos), None),
        ("Sugerencias de autocompletado", len(sugerencias), None),
        ("Preguntas PAA (SERP)", len(preguntas), None),
        ("Preguntas (Autocompletado)", len(preguntas_ac), None),
        ("Busquedas relacionadas", len(relacionadas), None),
        ("Keywords con Google Ads", con_ads, None),
        ("Keywords con Google Trends", con_trends, None),
    ]

    for label, value, row_fill in stats:
        row += 1
        cell_label = _aplicar_celda(ws, row, 1, label, font=FONT_LABEL, alignment=ALIGN_LEFT)
        cell_value = _aplicar_celda(ws, row, 2, value, font=FONT_BODY, alignment=ALIGN_LEFT)
        if row_fill:
            cell_label.fill = row_fill
            cell_value.fill = row_fill

    row += 1
    _aplicar_celda(ws, row, 1, "TOTAL DE RESULTADOS", font=FONT_BODY_BOLD, alignment=ALIGN_LEFT, fill=FILL_ACCENT_LIGHT)
    _aplicar_celda(ws, row, 2, total, font=FONT_STAT_VALUE, alignment=ALIGN_LEFT, fill=FILL_ACCENT_LIGHT)

    row += 3
    ws.merge_cells(f"A{row}:D{row}")
    _aplicar_celda(ws, row, 1, "DISTRIBUCION POR PRIORIDAD", font=FONT_HEADER, alignment=ALIGN_CENTER, fill=FILL_HEADER)
    for col in range(2, 5):
        ws.cell(row=row, column=col).border = BORDER
        ws.cell(row=row, column=col).fill = FILL_HEADER

    categorias_orden = ["Muy alta", "Alta", "Media", "Baja", "Muy baja"]
    conteo = {categoria: 0 for categoria in categorias_orden}
    for metrica in volumenes.values():
        conteo[_categorizar_score_texto(metrica.get("score", 0))] += 1

    row += 1
    _aplicar_celda(ws, row, 1, "Prioridad", font=FONT_HEADER, alignment=ALIGN_CENTER, fill=FILL_HEADER)
    ws.merge_cells(f"B{row}:D{row}")
    _aplicar_celda(ws, row, 2, "Cantidad", font=FONT_HEADER, alignment=ALIGN_CENTER, fill=FILL_HEADER)
    for col in range(3, 5):
        ws.cell(row=row, column=col).border = BORDER
        ws.cell(row=row, column=col).fill = FILL_HEADER

    for categoria in categorias_orden:
        row += 1
        cantidad = conteo[categoria]

        _aplicar_celda(ws, row, 1, categoria, font=FONT_BODY_BOLD, alignment=ALIGN_LEFT)
        ws.merge_cells(f"B{row}:D{row}")
        _aplicar_celda(ws, row, 2, cantidad, font=FONT_BODY, alignment=ALIGN_CENTER)
        for col in range(3, 5):
            ws.cell(row=row, column=col).border = BORDER

        ws.cell(row=row, column=1).fill = _fill_score({"Muy alta": 90, "Alta": 60, "Media": 35, "Baja": 20, "Muy baja": 5}[categoria])

    row += 3
    _aplicar_celda(ws, row, 1, "LECTURA EJECUTIVA", font=FONT_SECTION, alignment=ALIGN_LEFT)
    row += 1
    ws.merge_cells(f"A{row}:D{row + 4}")
    lectura = (
        "Este archivo sirve para priorizar contenido, detectar preguntas reales del mercado y ordenar "
        "la investigacion SEO. Cuando Google Ads no esta disponible, la lectura correcta es editorial: "
        "Muy alta/Alta indica temas que conviene validar primero; Media amplia cobertura; Baja/Muy baja "
        "apunta a nichos o temas secundarios. No usar este reporte para estimar ingresos, presupuestos "
        "de pauta o forecasting financiero."
    )
    cell = ws.cell(row=row, column=1, value=lectura)
    cell.font = Font(name="Century Gothic", size=9, color=_DARK)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    row += 6
    _aplicar_celda(ws, row, 1, "METODOLOGIA", font=FONT_SECTION, alignment=ALIGN_LEFT)
    row += 1
    ws.merge_cells(f"A{row}:D{row + 4}")
    nota = (
        "Este archivo no estima busquedas mensuales exactas ni muestra rangos inventados. "
        "Solo conserva datos trazables de Google: la fuente donde se detecto cada termino, "
        "su posicion dentro de esa fuente y, cuando esta disponible, metricas historicas reales "
        "de Google Ads y el interes relativo de Google Trends. El score sirve solo para priorizar "
        "temas dentro del reporte."
    )
    cell = ws.cell(row=row, column=1, value=nota)
    cell.font = Font(name="Century Gothic", size=9, color=_GRAY, italic=True)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 94.3
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 2.7
    ws.sheet_properties.tabColor = _ACCENT
    return ws


def exportar_excel(keyword: str, datos: Dict[str, List[str]]) -> str:
    """Exporta el reporte a Excel con trazabilidad real."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    if EXCEL_TEMPLATE_PATH and os.path.exists(EXCEL_TEMPLATE_PATH):
        wb = load_workbook(EXCEL_TEMPLATE_PATH)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    volumenes = datos.get("volumenes", {})
    language_code = datos.get("language_code", "es")

    _crear_hoja_resumen(wb, keyword, datos, volumenes)

    sugerencias = datos.get("sugerencias", [])
    if sugerencias:
        _crear_hoja_datos(wb, "Autocompletado", sugerencias, volumenes, "Sugerencia", FILL_HEADER, language_code)

    preguntas = datos.get("preguntas_paa", [])
    if preguntas:
        _crear_hoja_datos(wb, "Preguntas PAA", preguntas, volumenes, "Pregunta", FILL_HEADER_ALT, language_code)

    preguntas_ac = datos.get("preguntas_autocompletado", [])
    if preguntas_ac:
        _crear_hoja_datos(wb, "Preguntas Autocompletado", preguntas_ac, volumenes, "Pregunta", FILL_HEADER_ALT, language_code)

    relacionadas = datos.get("busquedas_relacionadas", [])
    if relacionadas:
        _crear_hoja_datos(wb, "Busquedas relacionadas", relacionadas, volumenes, "Busqueda relacionada", FILL_HEADER, language_code)

    nombre = generar_nombre_archivo(keyword, "xlsx")
    ruta = os.path.join(OUTPUT_DIR, nombre)
    wb.save(ruta)
    return ruta
